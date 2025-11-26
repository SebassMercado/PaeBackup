from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.core.paginator import Paginator
from datetime import datetime
import json


def listar_producciones(request):
    """
    Vista para listar producciones usando tu estructura actual
    """
    # Filtros
    estado_filter = request.GET.get('estado', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Query base con SQL directo - consulta simplificada
    with connection.cursor() as cursor:
        # Traer producción y, si existe, venta asociada para obtener
        # usuarios de registro y asignación usando la venta como fuente primaria.
        sql = """
            SELECT 
                p.id_proc, p.estado, p.fecha_hora, p.fecha_aceptacion,
                p.fecha_finalizacion, p.id_usu AS prod_id_usu, p.id_asignado AS prod_id_asignado,
                v.id_ven AS venta_id,
                v.id_usu AS venta_id_usu, v.id_asignado AS venta_id_asignado,
                COALESCE(u_v_reg.nombres, u_reg.nombres, '') AS usuario_registro_nombres,
                COALESCE(u_v_reg.apellidos, u_reg.apellidos, '') AS usuario_registro_apellidos,
                COALESCE(u_v_asig.nombres, u_asig.nombres, '') AS asignado_a_nombres,
                COALESCE(u_v_asig.apellidos, u_asig.apellidos, '') AS asignado_a_apellidos
            FROM produccion p
            LEFT JOIN venta_produccion vp ON vp.id_produccion = p.id_proc
            LEFT JOIN ventas v ON v.id_ven = vp.id_venta
            LEFT JOIN usuarios u_reg ON u_reg.id_usu = p.id_usu
            LEFT JOIN usuarios u_asig ON u_asig.id_usu = p.id_asignado
            LEFT JOIN usuarios u_v_reg ON u_v_reg.id_usu = v.id_usu
            LEFT JOIN usuarios u_v_asig ON u_v_asig.id_usu = v.id_asignado
            WHERE 1=1
        """
        params = []
        
        # Aplicar filtros
        if estado_filter:
            sql += " AND p.estado = %s"
            params.append(estado_filter)
        
        if fecha_inicio:
            sql += " AND DATE(p.fecha_hora) >= %s"
            params.append(fecha_inicio)
        
        if fecha_fin:
            sql += " AND DATE(p.fecha_hora) <= %s"
            params.append(fecha_fin)
        
        sql += " ORDER BY p.fecha_hora DESC"
        
        cursor.execute(sql, params)
        columns = [col[0] for col in cursor.description]
        raw_rows = cursor.fetchall()
        producciones_raw = []
        for row in raw_rows:
            data = dict(zip(columns, row))
            # Construir nombre completo priorizando origen venta si existe.
            usuario_registro = (f"{data['usuario_registro_nombres']} {data['usuario_registro_apellidos']}" ).strip()
            asignado_a = (f"{data['asignado_a_nombres']} {data['asignado_a_apellidos']}" ).strip()
            if not usuario_registro:
                usuario_registro = 'Sistema'
            if not asignado_a:
                asignado_a = 'Sin asignar'
            # Normalizar claves esperadas por template
            producciones_raw.append({
                'id_proc': data['id_proc'],
                'estado': data['estado'],
                'fecha_hora': data['fecha_hora'],
                'usuario_registro': usuario_registro,
                'asignado_a': asignado_a,
                'venta_id': data.get('venta_id') or None,
            })
    
    # Paginación
    paginator = Paginator(producciones_raw, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    with connection.cursor() as cursor:
        cursor.execute("SELECT COUNT(*) FROM produccion")
        total = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM produccion WHERE estado = 'Pendiente'")
        pendientes = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM produccion WHERE estado = 'Aceptada'")
        en_proceso = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM produccion WHERE estado = 'Finalizada'")
        completadas = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM produccion WHERE estado = 'Esperando insumos'")
        esperando = cursor.fetchone()[0]
    
    stats = {
        'total': total,
        'pendientes': pendientes,
        'en_proceso': en_proceso,
        'completadas': completadas,
        'esperando': esperando,
    }
    
    # Obtener usuarios para filtros
    with connection.cursor() as cursor:
        cursor.execute("SELECT id_usu FROM usuarios ORDER BY id_usu")
        usuarios = [{'id': row[0], 'nombre': f'Usuario {row[0]}'} for row in cursor.fetchall()]
    
    context = {
        'producciones': page_obj,
        'stats': stats,
        'usuarios': usuarios,
        'filtros': {
            'estado': estado_filter,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
        },
        'estados_choices': [
            ('Pendiente', 'Pendiente'),
            ('Aceptada', 'Aceptada'),
            ('Finalizada', 'Finalizada'),
            ('Esperando insumos', 'Esperando insumos'),
        ]
    }
    
    return render(request, 'produccion/index.html', context)


def nueva_produccion(request):
    """
    Vista para crear nueva producción
    """
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            id_usu = request.POST.get('id_usu', 1)  # Usuario que crea
            id_asignado = request.POST.get('id_asignado')  # Usuario asignado
            
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO produccion (estado, fecha_hora, id_usu, id_asignado)
                    VALUES (%s, %s, %s, %s)
                """, ['Pendiente', datetime.now(), id_usu, id_asignado])
            
            messages.success(request, 'Producción creada exitosamente')
            return redirect('produccion:listar')
            
        except Exception as e:
            messages.error(request, f'Error al crear la producción: {str(e)}')
    
    # Obtener usuarios para asignación
    with connection.cursor() as cursor:
        cursor.execute("SELECT id_usu FROM usuarios ORDER BY id_usu")
        usuarios = [{'id': row[0], 'nombre': f'Usuario {row[0]}'} for row in cursor.fetchall()]
    
    context = {'usuarios': usuarios}
    return render(request, 'produccion/nueva.html', context)


def cambiar_estado_produccion(request, produccion_id):
    """
    Cambia el estado de una producción
    """
    try:
        with connection.cursor() as cursor:
            # Obtener producción actual
            cursor.execute("SELECT estado FROM produccion WHERE id_proc = %s", [produccion_id])
            result = cursor.fetchone()
            
            if not result:
                messages.error(request, 'Producción no encontrada')
                return redirect('produccion:listar')
            
            estado_actual = result[0]
            fecha_actual = datetime.now()
            
            if estado_actual == 'Pendiente':
                # Pendiente → Aceptada
                cursor.execute("""
                    UPDATE produccion 
                    SET estado = 'Aceptada', fecha_aceptacion = %s 
                    WHERE id_proc = %s
                """, [fecha_actual, produccion_id])
                messages.success(request, f'Producción #{produccion_id} aceptada')
                
            elif estado_actual == 'Aceptada':
                # Aceptada → Finalizada
                cursor.execute("""
                    UPDATE produccion 
                    SET estado = 'Finalizada', fecha_finalizacion = %s 
                    WHERE id_proc = %s
                """, [fecha_actual, produccion_id])
                messages.success(request, f'Producción #{produccion_id} finalizada')
                
            elif estado_actual == 'Esperando insumos':
                # Esperando insumos → Finalizada
                cursor.execute("""
                    UPDATE produccion 
                    SET estado = 'Finalizada', fecha_finalizacion = %s 
                    WHERE id_proc = %s
                """, [fecha_actual, produccion_id])
                messages.success(request, f'Producción #{produccion_id} finalizada')
                
            else:
                messages.info(request, 'La producción ya está finalizada')
                
    except Exception as e:
        messages.error(request, f'Error al cambiar estado: {str(e)}')
    
    return redirect('produccion:listar')


def _transicionar_estado_produccion(produccion_id):
    """Lógica reutilizable para cambiar estado secuencial de producción.
    Retorna (success: bool, mensaje: str, nuevo_estado: str|None)."""
    from datetime import datetime
    fecha_actual = datetime.now()
    with connection.cursor() as cursor:
        cursor.execute("SELECT estado FROM produccion WHERE id_proc = %s", [produccion_id])
        result = cursor.fetchone()
        if not result:
            return False, 'Producción no encontrada', None
        estado_actual = result[0]
        if estado_actual == 'Pendiente':
            cursor.execute("""
                UPDATE produccion
                SET estado = 'Aceptada', fecha_aceptacion = %s
                WHERE id_proc = %s
            """, [fecha_actual, produccion_id])
            return True, f'Producción #{produccion_id} aceptada', 'Aceptada'
        if estado_actual == 'Aceptada':
            # Para finalizar se espera que API reciba observacion/descontar; esta función secuencial no los procesa.
            cursor.execute("""
                UPDATE produccion
                SET estado = 'Finalizada', fecha_finalizacion = %s
                WHERE id_proc = %s AND estado='Aceptada'
            """, [fecha_actual, produccion_id])
            return True, f'Producción #{produccion_id} finalizada', 'Finalizada'
        if estado_actual == 'Esperando insumos':
            # No finalizar aquí para forzar verificación con descuento real
            return False, 'Faltan insumos, use reintentar finalización tras reponer stock', 'Esperando insumos'
        return False, 'La producción ya está finalizada', estado_actual


def api_cambiar_estado_produccion(request, produccion_id):
    """Endpoint API (JSON) para cambiar estado vía AJAX."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'})
    try:
        # Opcionalmente leer acción enviada aunque la lógica sea secuencial
        data = {}
        if request.body:
            try:
                data = json.loads(request.body.decode('utf-8'))
            except json.JSONDecodeError:
                pass
        accion = data.get('accion')
        observacion = data.get('observacion')

        if accion == 'finalizar':
            # Finalizar: siempre descuenta insumos; observación opcional.
            ok, msg, estado = _finalizar_produccion_con_insumos(produccion_id, observacion)
            return JsonResponse({'success': ok, 'message': msg, 'estado': estado})
        else:
            success, mensaje, nuevo_estado = _transicionar_estado_produccion(produccion_id)
            return JsonResponse({'success': success, 'message': mensaje, 'estado': nuevo_estado})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {e}'})


def _finalizar_produccion_con_insumos(produccion_id, observacion=None):
    """Finaliza producción aceptada. Descuenta insumos automáticamente usando lotes por vencimiento (FIFO).
    Retorna (success, mensaje, estado)."""
    from decimal import Decimal
    fecha_actual = datetime.now()
    with connection.cursor() as cursor:
        cursor.execute("SELECT estado FROM produccion WHERE id_proc=%s", [produccion_id])
        row = cursor.fetchone()
        if not row:
            return False, 'Producción no encontrada', None
        if row[0] not in ('Aceptada', 'Esperando insumos'):
            return False, f"Estado actual '{row[0]}' no permite finalizar", row[0]

        # Verificación previa de disponibilidad de insumos para TODAS las recetas
        cursor.execute("""
            SELECT ri.id_ins,
                   SUM(ri.cantidad * pr.cantidad) AS requerido_total,
                   i.stock_actual,
                   i.nombre
            FROM produccion_recetas pr
            JOIN receta_insumos ri ON pr.id_rec = ri.id_rec
            JOIN insumos i ON ri.id_ins = i.id_ins
            WHERE pr.id_produccion = %s
            GROUP BY ri.id_ins, i.stock_actual, i.nombre
        """, [produccion_id])
        faltantes = []
        disponibilidad_map = []
        for id_ins, requerido_total, stock_actual, nombre in cursor.fetchall():
            requerido_dec = Decimal(str(requerido_total or 0))
            stock_dec = Decimal(str(stock_actual or 0))
            if stock_dec < requerido_dec or stock_dec <= 0:
                faltantes.append({'id_ins': id_ins, 'nombre': nombre, 'requerido': str(requerido_dec), 'disponible': str(stock_dec)})
            disponibilidad_map.append((id_ins, requerido_dec, stock_dec))

        if faltantes:
            # Cambiar a 'Esperando insumos' y registrar observación detallando faltantes (sin descontar nada)
            detalle_faltantes = ", ".join([f"{f['nombre']} (Req: {f['requerido']}, Disp: {f['disponible']})" for f in faltantes])
            observ_text = (observacion or '').strip()[:1000] if observacion else ''
            observ_final = (observ_text + (' | ' if observ_text else '') + 'Faltan insumos: ' + detalle_faltantes)[:2000]
            cursor.execute("""
                UPDATE produccion SET estado='Esperando insumos', observacion=%s
                WHERE id_proc=%s AND estado='Aceptada'
            """, [observ_final, produccion_id])
            return False, f"Producción #{produccion_id} en 'Esperando insumos'. Faltan: {detalle_faltantes}", 'Esperando insumos'

        observ_text = (observacion or '').strip()[:2000] if observacion else None
        cursor.execute("""
            UPDATE produccion SET estado='Finalizada', fecha_finalizacion=%s, observacion=%s
            WHERE id_proc=%s AND estado='Aceptada'
        """, [fecha_actual, observ_text, produccion_id])

        detalles_descontados = []
        # Obtener recetas asociadas a la producción
        cursor.execute("SELECT id_rec, cantidad FROM produccion_recetas WHERE id_produccion=%s", [produccion_id])
        recetas = cursor.fetchall()
        for id_rec, cant_prod in recetas:
            # Ingredientes de la receta
            cursor.execute("SELECT id_ins, cantidad FROM receta_insumos WHERE id_rec=%s", [id_rec])
            ingredientes = cursor.fetchall()
            for id_ins, cant_por_unidad in ingredientes:
                requerido_total = Decimal(str(cant_por_unidad)) * Decimal(str(cant_prod))
                # Consumir de lotes activos ordenados por vencimiento (los más próximos primero, NULLs al final)
                cursor.execute("""
                    SELECT id_detalle, cantidad, fecha_vencimiento FROM detalle_insumo
                    WHERE id_ins=%s AND estado='Activo'
                    ORDER BY (fecha_vencimiento IS NULL), fecha_vencimiento ASC, fecha_ingreso ASC
                """, [id_ins])
                lotes = cursor.fetchall()
                restante = requerido_total
                for id_detalle, cant_lote, fecha_venc in lotes:
                    cant_lote_dec = Decimal(str(cant_lote))
                    if cant_lote_dec <= 0:
                        continue
                    if restante <= 0:
                        break
                    if cant_lote_dec >= restante:
                        nuevo_lote = cant_lote_dec - restante
                        cursor.execute("UPDATE detalle_insumo SET cantidad=%s, estado=%s WHERE id_detalle=%s", [nuevo_lote, 'Agotado' if nuevo_lote == 0 else 'Activo', id_detalle])
                        restante = Decimal('0')
                    else:
                        # Agotar lote
                        restante -= cant_lote_dec
                        cursor.execute("UPDATE detalle_insumo SET cantidad=0, estado='Agotado' WHERE id_detalle=%s", [id_detalle])
                # Actualizar stock general del insumo
                cursor.execute("UPDATE insumos SET stock_actual = GREATEST(stock_actual - %s, 0) WHERE id_ins=%s", [requerido_total, id_ins])
                # Re-sincronizar estado usando util centralizada
                try:
                    from modelos.insumos.utils import sincronizar_estado_insumo
                    sincronizar_estado_insumo(id_ins)
                except Exception as e:
                    print(f"[WARN] No se pudo sincronizar estado insumo {id_ins}: {e}")
                detalles_descontados.append({'insumo': id_ins, 'cantidad': str(requerido_total), 'restante_no_cubierto': str(restante)})
        # Actualizar estado de venta asociada (si existe) a 'Completada'
        try:
            cursor.execute("SELECT id_venta FROM venta_produccion WHERE id_produccion=%s", [produccion_id])
            venta_row = cursor.fetchone()
            if venta_row and venta_row[0]:
                id_venta = venta_row[0]
                cursor.execute("UPDATE ventas SET estado='Completada' WHERE id_ven=%s", [id_venta])
        except Exception as e:
            print(f"[WARN] No se pudo actualizar venta asociada a producción {produccion_id}: {e}")

        msg = f"Producción #{produccion_id} finalizada" + (" con observación" if observ_text else "") + f" | Insumos descontados: {len(detalles_descontados)} insumos"
        return True, msg, 'Finalizada'


def eliminar_produccion(request, produccion_id):
    """
    Elimina una producción
    """
    try:
        with connection.cursor() as cursor:
            # Verificar que la producción esté en estado eliminable
            cursor.execute("SELECT estado FROM produccion WHERE id_proc = %s", [produccion_id])
            result = cursor.fetchone()
            
            if not result:
                messages.error(request, 'Producción no encontrada')
                return redirect('produccion:listar')
            
            estado = result[0]
            if estado not in ['Pendiente', 'Esperando insumos']:
                messages.error(request, 'Solo se pueden eliminar producciones pendientes o esperando insumos')
                return redirect('produccion:listar')
            
            # Eliminar la producción
            cursor.execute("DELETE FROM produccion WHERE id_proc = %s", [produccion_id])
            messages.success(request, f'Producción #{produccion_id} eliminada correctamente')
            
    except Exception as e:
        messages.error(request, f'Error al eliminar la producción: {str(e)}')
    
    return redirect('produccion:listar')


def detalle_produccion(request, produccion_id):
    """
    Vista para mostrar el detalle de una producción
    """
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT p.id_proc, p.estado, p.fecha_hora, p.fecha_aceptacion, 
                   p.fecha_finalizacion, p.id_usu, p.id_asignado, p.observacion
            FROM produccion p
            WHERE p.id_proc = %s
        """, [produccion_id])
        
        result = cursor.fetchone()
        if not result:
            messages.error(request, 'Producción no encontrada')
            return redirect('produccion:listar')
        
        columns = ['id_proc', 'estado', 'fecha_hora', 'fecha_aceptacion', 
              'fecha_finalizacion', 'id_usu', 'id_asignado', 'observacion']
        produccion = dict(zip(columns, result))

        # Obtener nombres reales (nombres + apellidos) si existen
        creador_nombres = creador_apellidos = asignado_nombres = asignado_apellidos = ''
        if produccion['id_usu']:
            cursor.execute("SELECT nombres, apellidos FROM usuarios WHERE id_usu=%s", [produccion['id_usu']])
            row_creador = cursor.fetchone()
            if row_creador:
                creador_nombres, creador_apellidos = row_creador
        if produccion['id_asignado']:
            cursor.execute("SELECT nombres, apellidos FROM usuarios WHERE id_usu=%s", [produccion['id_asignado']])
            row_asig = cursor.fetchone()
            if row_asig:
                asignado_nombres, asignado_apellidos = row_asig

        produccion['usuario_nombre'] = (f"{creador_nombres} {creador_apellidos}".strip() or (f'Usuario {produccion["id_usu"]}' if produccion['id_usu'] else 'Sistema'))
        produccion['asignado_nombre'] = (f"{asignado_nombres} {asignado_apellidos}".strip() or (f'Usuario {produccion["id_asignado"]}' if produccion['id_asignado'] else 'Sin asignar'))
        produccion['nombre_completo_creador'] = produccion['usuario_nombre']
        produccion['nombre_completo_asignado'] = produccion['asignado_nombre']
        
        # Verificar si esta producción fue generada desde una venta
        cursor.execute("""
            SELECT vp.id_venta 
            FROM venta_produccion vp 
            WHERE vp.id_produccion = %s
        """, [produccion_id])
        
        venta_origen_result = cursor.fetchone()
        produccion['venta_origen'] = venta_origen_result[0] if venta_origen_result else None
    
    context = {
        'produccion': produccion,
    }
    
    return render(request, 'produccion/detalle.html', context)


def obtener_fechas_produccion(request, produccion_id):
    """
    API para obtener las fechas de una producción
    """
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT fecha_hora, fecha_aceptacion, fecha_finalizacion
                FROM produccion WHERE id_proc = %s
            """, [produccion_id])
            
            result = cursor.fetchone()
            if not result:
                return JsonResponse({'success': False, 'message': 'Producción no encontrada'})
            
            fecha_hora, fecha_aceptacion, fecha_finalizacion = result
            
            return JsonResponse({
                'success': True,
                'fecha_hora': fecha_hora.strftime('%d/%m/%Y %H:%M:%S') if fecha_hora else None,
                'fecha_aceptacion': fecha_aceptacion.strftime('%d/%m/%Y %H:%M:%S') if fecha_aceptacion else None,
                'fecha_finalizacion': fecha_finalizacion.strftime('%d/%m/%Y %H:%M:%S') if fecha_finalizacion else None,
            })
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


def exportar_excel(request):
    """
    Exporta las producciones a Excel
    """
    try:
        from openpyxl import Workbook
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill
        
        # Crear libro de trabajo
        wb = Workbook()
        ws = wb.active
        ws.title = "Producciones"
        
        # Encabezados
        headers = ['ID', 'Estado', 'Fecha Creación', 'Fecha Aceptación', 'Fecha Finalización', 'Usuario', 'Asignado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Obtener datos
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT p.id_proc, p.estado, p.fecha_hora, p.fecha_aceptacion, 
                       p.fecha_finalizacion, p.id_usu, p.id_asignado
                FROM produccion p
                ORDER BY p.fecha_hora DESC
            """)
            
            # Datos
            for row_num, prod in enumerate(cursor.fetchall(), 2):
                ws.cell(row=row_num, column=1, value=prod[0])  # ID
                ws.cell(row=row_num, column=2, value=prod[1])  # Estado
                ws.cell(row=row_num, column=3, value=prod[2].strftime('%d/%m/%Y %H:%M:%S') if prod[2] else '')  # Fecha creación
                ws.cell(row=row_num, column=4, value=prod[3].strftime('%d/%m/%Y %H:%M:%S') if prod[3] else '')  # Fecha aceptación
                ws.cell(row=row_num, column=5, value=prod[4].strftime('%d/%m/%Y %H:%M:%S') if prod[4] else '')  # Fecha finalización
                ws.cell(row=row_num, column=6, value=f'Usuario {prod[5]}' if prod[5] else 'Sistema')  # Usuario
                ws.cell(row=row_num, column=7, value=f'Usuario {prod[6]}' if prod[6] else 'Sin asignar')  # Asignado
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="producciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Error al exportar Excel: {str(e)}')
        return redirect('produccion:listar')


# ==================== STUB: EXPORTAR PDF PRODUCCIONES ====================
def exportar_pdf(request):
    """Stub para generación de reporte PDF de producciones.
    Implementación futura: listar producciones con estados y fechas.
    Ahora solo muestra un mensaje y redirige."""
    messages.info(request, "ℹ️ Reporte PDF de producciones aún no implementado.")
    return redirect('produccion:index')


# ==================== STUB: MIGRAR PRODUCCIONES ====================
def migrar_producciones(request):
    """Stub para migración de producciones (import/export masivo).
    Aquí podría manejar cargas desde CSV/Excel o APIs externas."""
    messages.info(request, "ℹ️ Módulo de migración de producciones aún no implementado.")
    return redirect('produccion:index')


def recetas_produccion(request, produccion_id):
    """
    Vista para mostrar las recetas asociadas a una producción
    """
    # Obtener información de la producción con fechas de trazabilidad
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT p.id_proc, p.estado, p.fecha_hora, p.fecha_aceptacion, 
                   p.fecha_finalizacion, p.id_usu, p.id_asignado, p.observacion
            FROM produccion p
            WHERE p.id_proc = %s
        """, [produccion_id])
        
        result = cursor.fetchone()
        if not result:
            messages.error(request, 'Producción no encontrada')
            return redirect('produccion:listar')
        
        produccion = {
            'id_proc': result[0],
            'estado': result[1],
            'fecha_hora': result[2],
            'fecha_aceptacion': result[3],
            'fecha_finalizacion': result[4],
            'id_usu': result[5],
            'id_asignado': result[6],
            'observacion': result[7],
        }
        # Nombres reales
        creador_nombres = creador_apellidos = asignado_nombres = asignado_apellidos = ''
        if produccion['id_usu']:
            cursor.execute("SELECT nombres, apellidos FROM usuarios WHERE id_usu=%s", [produccion['id_usu']])
            row_creador = cursor.fetchone()
            if row_creador:
                creador_nombres, creador_apellidos = row_creador
        if produccion['id_asignado']:
            cursor.execute("SELECT nombres, apellidos FROM usuarios WHERE id_usu=%s", [produccion['id_asignado']])
            row_asig = cursor.fetchone()
            if row_asig:
                asignado_nombres, asignado_apellidos = row_asig
        produccion['usuario_nombre'] = (f"{creador_nombres} {creador_apellidos}".strip() or (f"Usuario {produccion['id_usu']}" if produccion['id_usu'] else 'Sistema'))
        produccion['asignado_nombre'] = (f"{asignado_nombres} {asignado_apellidos}".strip() or (f"Usuario {produccion['id_asignado']}" if produccion['id_asignado'] else 'Sin asignar'))
        produccion['nombre_completo_creador'] = produccion['usuario_nombre']
        produccion['nombre_completo_asignado'] = produccion['asignado_nombre']
    
    # Obtener recetas asociadas desde la tabla produccion_recetas
    recetas_produccion = []
    
    with connection.cursor() as cursor:
        try:
            # Consultar directamente la tabla produccion_recetas
            cursor.execute("""
                SELECT pr.id_detalle, pr.id_rec, pr.cantidad, r.nombre
                FROM produccion_recetas pr
                LEFT JOIN recetas r ON pr.id_rec = r.id_rec
                WHERE pr.id_produccion = %s
                ORDER BY pr.id_detalle
            """, [produccion_id])
            
            for row in cursor.fetchall():
                recetas_produccion.append({
                    'id_detalle': row[0],
                    'id_rec': row[1],
                    'cantidad': row[2],
                    'receta': {
                        'id_rec': row[1],
                        'nombre': row[3] if row[3] else f'Receta #{row[1]}'
                    }
                })
                
            print(f"🍳 Recetas encontradas para producción #{produccion_id}: {len(recetas_produccion)}")
                    
        except Exception as e:
            print(f"⚠️ Error al obtener recetas de producción: {e}")
    
    # Obtener recetas disponibles para agregar
    recetas_disponibles = []
    
    with connection.cursor() as cursor:
        try:
            # Obtener todas las recetas disponibles que no estén ya en esta producción
            cursor.execute("""
                SELECT r.id_rec, r.nombre 
                FROM recetas r
                WHERE r.id_rec NOT IN (
                    SELECT pr.id_rec 
                    FROM produccion_recetas pr 
                    WHERE pr.id_produccion = %s
                )
                ORDER BY r.nombre
            """, [produccion_id])
            
            recetas_disponibles = [{'id_rec': row[0], 'nombre': row[1]} for row in cursor.fetchall()]
            
        except Exception as e:
            print(f"⚠️ Error al obtener recetas disponibles: {e}")
            # Fallback: obtener todas las recetas
            try:
                cursor.execute("SELECT id_rec, nombre FROM recetas ORDER BY nombre")
                recetas_disponibles = [{'id_rec': row[0], 'nombre': row[1]} for row in cursor.fetchall()]
            except:
                recetas_disponibles = []
    
    context = {
        'produccion': produccion,
        'recetas_produccion': recetas_produccion,
        'recetas_disponibles': recetas_disponibles,
        'receta_seleccionada': None,
        'ingredientes_receta': [],
    }
    
    return render(request, 'produccion/recetas.html', context)


def obtener_ingredientes_receta(request, receta_id):
    """
    API para obtener los ingredientes de una receta con su estado
    """
    if request.method == 'GET':
        try:
            with connection.cursor() as cursor:
                # Obtener ingredientes de la receta con estado de insumos
                cursor.execute("""
                    SELECT 
                        ri.id_rec_ins,
                        ri.cantidad,
                        ri.unidad,
                        ri.estado as estado_receta_insumo,
                        i.id_ins,
                        i.nombre,
                        i.unidad_medida,
                        i.stock_actual,
                        i.stock_min,
                        i.estado as estado_insumo_db,
                        CASE 
                            WHEN i.stock_actual <= 0 THEN 'Agotado'
                            WHEN i.stock_actual < i.stock_min THEN 'Insuficiente'
                            ELSE 'Activo'
                        END as estado_insumo_calculado,
                        CASE 
                            WHEN ri.estado = 'Agotado' OR i.stock_actual <= 0 THEN 'Agotado'
                            WHEN i.stock_actual >= ri.cantidad AND i.stock_actual >= i.stock_min THEN 'Disponible'
                            WHEN i.stock_actual > 0 THEN 'Insuficiente'
                            ELSE 'Agotado'
                        END as disponibilidad
                    FROM receta_insumos ri
                    INNER JOIN insumos i ON ri.id_ins = i.id_ins
                    WHERE ri.id_rec = %s
                    ORDER BY i.nombre
                """, [receta_id])
                
                ingredientes = []
                tiene_ingrediente_agotado = False
                
                for row in cursor.fetchall():
                    estado_final = row[11]  # disponibilidad
                    if estado_final in ['Agotado', 'Insuficiente']:
                        tiene_ingrediente_agotado = True
                    
                    # Usar la unidad de medida de la receta (más específica)
                    unidad_medida = row[2] if row[2] else row[6]  # ri.unidad o i.unidad_medida
                    
                    ingredientes.append({
                        'id_rec_ins': row[0],
                        'cantidad': float(row[1]),
                        'unidad_receta': row[2],  # unidad específica de la receta
                        'estado_en_receta': row[3],  # estado en tabla receta_insumos
                        'insumo': {
                            'id_ins': row[4],
                            'nombre': row[5],
                            'unidad_medida': unidad_medida,
                            'stock_actual': float(row[7]) if row[7] else 0,
                            'stock_min': float(row[8]) if row[8] else 0,
                            'estado_db': row[9],  # estado desde BD (Activo/Inactivo/Stock insuficiente)
                            'estado': row[10]  # estado calculado (Agotado/Stock bajo/Activo)
                        },
                        'estado': estado_final,  # disponibilidad final
                        'disponible': estado_final == 'Disponible'
                    })
                
                # Obtener nombre de la receta
                cursor.execute("SELECT nombre FROM recetas WHERE id_rec = %s", [receta_id])
                receta_result = cursor.fetchone()
                receta_nombre = receta_result[0] if receta_result else f'Receta #{receta_id}'
                
                # Determinar estado de la receta basado en ingredientes
                estado_receta = 'Inactiva' if tiene_ingrediente_agotado else 'Activa'
                
                return JsonResponse({
                    'success': True,
                    'receta': {
                        'id': receta_id,
                        'nombre': receta_nombre,
                        'estado': estado_receta
                    },
                    'ingredientes': ingredientes,
                    'total_ingredientes': len(ingredientes),
                    'ingredientes_disponibles': len([i for i in ingredientes if i['disponible']]),
                    'ingredientes_agotados': len([i for i in ingredientes if not i['disponible']])
                })
                
        except Exception as e:
            print(f"Error al obtener ingredientes de receta {receta_id}: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


def agregar_receta_produccion(request, produccion_id):
    """
    Agrega una receta a una producción
    """
    if request.method == 'POST':
        try:
            receta_id = request.POST.get('receta_id')
            cantidad = request.POST.get('cantidad', 1)
            
            if not receta_id or not cantidad:
                return JsonResponse({'success': False, 'message': 'Datos incompletos'})
            
            with connection.cursor() as cursor:
                # Verificar que la receta no esté ya asociada
                cursor.execute("""
                    SELECT COUNT(*) FROM produccion_recetas 
                    WHERE id_produccion = %s AND id_rec = %s
                """, [produccion_id, receta_id])
                
                if cursor.fetchone()[0] > 0:
                    return JsonResponse({'success': False, 'message': 'Esta receta ya está asociada a la producción'})
                
                # Agregar la receta
                cursor.execute("""
                    INSERT INTO produccion_recetas (id_produccion, id_rec, cantidad)
                    VALUES (%s, %s, %s)
                """, [produccion_id, receta_id, cantidad])
                
                # Obtener el nombre de la receta para el mensaje
                cursor.execute("SELECT nombre FROM recetas WHERE id_rec = %s", [receta_id])
                nombre_receta = cursor.fetchone()
                nombre_receta = nombre_receta[0] if nombre_receta else f'Receta #{receta_id}'
                
                return JsonResponse({
                    'success': True, 
                    'message': f'Receta "{nombre_receta}" agregada exitosamente'
                })
                
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'})


def actualizar_cantidad_receta(request, detalle_id):
    """
    Actualiza la cantidad de una receta en la producción
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            nueva_cantidad = data.get('cantidad')
            
            if not nueva_cantidad or nueva_cantidad <= 0:
                return JsonResponse({'success': False, 'message': 'Cantidad inválida'})
            
            with connection.cursor() as cursor:
                cursor.execute("""
                    UPDATE produccion_recetas 
                    SET cantidad = %s 
                    WHERE id_detalle = %s
                """, [nueva_cantidad, detalle_id])
                
                if cursor.rowcount > 0:
                    return JsonResponse({
                        'success': True, 
                        'message': f'Cantidad actualizada a {nueva_cantidad}'
                    })
                else:
                    return JsonResponse({'success': False, 'message': 'Registro no encontrado'})
                
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'})


def eliminar_receta_produccion(request, detalle_id):
    """
    Elimina una receta de la producción
    """
    if request.method == 'POST':
        try:
            with connection.cursor() as cursor:
                # Obtener información antes de eliminar
                cursor.execute("""
                    SELECT pr.id_rec, r.nombre 
                    FROM produccion_recetas pr
                    LEFT JOIN recetas r ON pr.id_rec = r.id_rec
                    WHERE pr.id_detalle = %s
                """, [detalle_id])
                
                result = cursor.fetchone()
                if not result:
                    return JsonResponse({'success': False, 'message': 'Registro no encontrado'})
                
                nombre_receta = result[1] if result[1] else f'Receta #{result[0]}'
                
                # Eliminar el registro
                cursor.execute("DELETE FROM produccion_recetas WHERE id_detalle = %s", [detalle_id])
                
                return JsonResponse({
                    'success': True, 
                    'message': f'Receta "{nombre_receta}" eliminada de la producción'
                })
                
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'})