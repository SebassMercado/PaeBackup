from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db.models import Q, Sum, Count, Avg, F, Min, Max
from django.core.paginator import Paginator
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.db import transaction
from decimal import Decimal
import json
from datetime import datetime, date

from .models import Produccion, ProduccionReceta
from modelos.usuarios.models import Usuario
from modelos.recetas.models import Receta, RecetaInsumo
from modelos.insumos.models import Insumo, DetalleInsumo
from modelos.ventas.models import Venta, VentaProduccion


# @login_required  # Temporalmente comentado para pruebas
def listar_producciones(request):
    """
    Vista para listar producciones
    Equivalente a: produccionBean.java - listar()
    """
    # Obtener filtros
    filtro_estado = request.GET.get('estado', '')
    filtro_fecha = request.GET.get('fecha', '')
    filtro_empleado = request.GET.get('empleado', '')
    filtro_receta = request.GET.get('receta', '')
    
    # Obtener usuario de sesión (simulado por ahora)
    usuario_actual = Usuario.objects.first()  # TODO: implementar sesión real
    
    # Query base - obtener todas las producciones por ahora
    producciones = Produccion.objects.all()
    
    # Aplicar filtros
    if filtro_estado:
        producciones = producciones.filter(estado=filtro_estado)
    
    if filtro_fecha:
        try:
            fecha_filtro = datetime.strptime(filtro_fecha, '%Y-%m-%d').date()
            producciones = producciones.filter(fecha_hora__date=fecha_filtro)
        except ValueError:
            pass
    
    if filtro_empleado and filtro_empleado.isdigit():
        producciones = producciones.filter(usuario_asignado_id=int(filtro_empleado))
    
    if filtro_receta and filtro_receta.isdigit():
        producciones = producciones.filter(receta_id=int(filtro_receta))
    
    # Ordenar por fecha descendente
    producciones = producciones.order_by('-fecha_hora')
    
    # Paginación
    paginator = Paginator(producciones, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener datos para filtros
    empleados_produccion = Usuario.objects.filter(
        role='P'  # Empleados de producción
    ).order_by('nombre')
    
    # Estadísticas rápidas
    stats = {
        'total': producciones.count(),
        'pendientes': producciones.filter(estado='Pendiente').count(),
        'en_proceso': producciones.filter(estado='En Proceso').count(),
        'finalizadas': producciones.filter(estado='Finalizada').count(),
    }
    
    context = {
        'producciones': producciones,
        'empleados_produccion': empleados_produccion,
        'stats': stats,
        'usuario_actual': usuario_actual,
    }
    
    return render(request, 'produccion/index.html', context)


# @login_required
def nueva_produccion(request):
    """
    Vista para crear nueva producción
    Equivalente a: produccionBean.java - guardar()
    """
    if request.method == 'GET':
        # Obtener datos para el formulario
        empleados_produccion = Usuario.objects.filter(
            rol='EP', estado='Activo'
        ).order_by('nombres')
        
        recetas_activas = Receta.objects.filter(
            estado='Activa'
        ).order_by('nombre')
        
        context = {
            'empleados_produccion': empleados_produccion,
            'recetas_activas': recetas_activas,
        }
        
        return render(request, 'produccion/nueva.html', context)
    
    # POST - procesar formulario
    return procesar_nueva_produccion(request)


@transaction.atomic
def procesar_nueva_produccion(request):
    """
    Procesa la creación de una nueva producción
    Equivalente a: produccionBean.java - guardar()
    """
    try:
        # Obtener datos del formulario
        receta_id = request.POST.get('id_receta')
        cantidad = request.POST.get('cantidad')
        empleado_id = request.POST.get('id_empleado')
        observaciones = request.POST.get('observaciones', '')
        
        # Validaciones básicas
        if not receta_id or not cantidad or not empleado_id:
            messages.error(request, 'Todos los campos son obligatorios')
            return redirect('produccion:nueva')
        
        try:
            cantidad = Decimal(cantidad)
            if cantidad <= 0:
                messages.error(request, 'La cantidad debe ser mayor a cero')
                return redirect('produccion:nueva')
        except (ValueError, TypeError):
            messages.error(request, 'Cantidad inválida')
            return redirect('produccion:nueva')
        
        # Obtener objetos relacionados
        receta = get_object_or_404(Receta, id_rec=receta_id)
        empleado = get_object_or_404(Usuario, id_usu=empleado_id, rol='EP')
        usuario_actual = Usuario.objects.first()  # TODO: implementar sesión real
        
        # Validar stock antes de crear la producción
        tiene_stock, mensaje_stock = validar_stock_produccion(receta, cantidad)
        
        # Crear la producción
        produccion = Produccion.objects.create(
            receta=receta,
            cantidad=cantidad,
            usuario_creador=usuario_actual,
            usuario_asignado=empleado,
            estado='Pendiente',
            fecha_hora=timezone.now()
        )
        
        # Si no hay stock suficiente, cambiar estado
        if not tiene_stock:
            produccion.estado = 'Esperando insumos'
            produccion.save()
            messages.warning(request, f'Producción creada pero en estado "Esperando insumos": {mensaje_stock}')
        else:
            messages.success(request, f'Producción #{produccion.id_proc} creada exitosamente')
        
        return redirect('produccion:index')
        
    except Exception as e:
        messages.error(request, f'Error al crear la producción: {str(e)}')
        return redirect('produccion:nueva')


def cambiar_estado_produccion(request, produccion_id):
    """
    Cambia el estado de una producción siguiendo el ciclo
    Equivalente a: produccionBean.java - cambiarEstadoCiclo()
    """
    try:
        produccion = get_object_or_404(Produccion, id_proc=produccion_id)
        fecha_actual = timezone.now()
        
        if produccion.estado == 'Pendiente':
            # Pendiente → Aceptada
            produccion.estado = 'En Proceso'
            produccion.fecha_aceptacion = fecha_actual
            produccion.save()
            
            messages.success(request, f'Producción #{produccion.id_proc} aceptada y puesta en proceso')
        
        elif produccion.estado == 'En Proceso':
            # En Proceso → Finalizada (con validación de stock)
            tiene_stock, mensaje_stock = validar_stock_produccion(produccion.receta, produccion.cantidad)
            
            if not tiene_stock:
                # Si no hay stock → Esperando insumos
                produccion.estado = 'Esperando insumos'
                produccion.save()
                messages.warning(request, f'Stock insuficiente. Estado cambiado a "Esperando insumos": {mensaje_stock}')
            else:
                # Si hay stock → Finalizar
                produccion.estado = 'Finalizada'
                produccion.fecha_finalizacion = fecha_actual
                produccion.save()
                
                # Descontar stock de insumos
                descontar_stock_produccion(produccion)
                
                # Marcar venta como completada si existe
                marcar_venta_completada(produccion)
                
                messages.success(request, f'Producción #{produccion.id_proc} finalizada exitosamente')
        
        elif produccion.estado == 'Esperando insumos':
            # Esperando insumos → intentar finalizar
            tiene_stock, mensaje_stock = validar_stock_produccion(produccion.receta, produccion.cantidad)
            
            if tiene_stock:
                produccion.estado = 'Finalizada'
                produccion.fecha_finalizacion = fecha_actual
                produccion.save()
                
                # Descontar stock
                descontar_stock_produccion(produccion)
                marcar_venta_completada(produccion)
                
                messages.success(request, f'Producción #{produccion.id_proc} finalizada. Stock disponible')
            else:
                messages.warning(request, f'Todavía no hay stock suficiente: {mensaje_stock}')
        
        else:
            messages.info(request, 'La producción ya está finalizada o no se puede cambiar su estado')
    
    except Exception as e:
        messages.error(request, f'Error al cambiar estado: {str(e)}')
    
    return redirect('produccion:index')


def detalle_produccion(request, produccion_id):
    """
    Vista para mostrar el detalle de una producción
    Equivalente a: produccionBean.java - verFechas()
    """
    produccion = get_object_or_404(Produccion, id_proc=produccion_id)
    
    # Obtener recetas asociadas (si hay tabla de relación)
    recetas_produccion = ProduccionReceta.objects.filter(
        produccion=produccion
    ).select_related('receta')
    
    # Obtener información de ingredientes necesarios
    ingredientes_necesarios = []
    if produccion.receta:
        ingredientes = RecetaInsumo.objects.filter(
            receta=produccion.receta
        ).select_related('insumo')
        
        for ingrediente in ingredientes:
            cantidad_necesaria = ingrediente.cantidad * produccion.cantidad
            stock_disponible = ingrediente.insumo.stock_actual
            suficiente = stock_disponible >= cantidad_necesaria
            
            ingredientes_necesarios.append({
                'insumo': ingrediente.insumo,
                'cantidad_unitaria': ingrediente.cantidad,
                'cantidad_necesaria': cantidad_necesaria,
                'stock_disponible': stock_disponible,
                'suficiente': suficiente,
                'faltante': max(0, cantidad_necesaria - stock_disponible),
            })
    
    # Calcular información adicional
    tiempo_transcurrido = None
    tiempo_produccion = None
    
    if produccion.fecha_hora:
        tiempo_transcurrido = timezone.now() - produccion.fecha_hora
    
    if produccion.fecha_finalizacion and produccion.fecha_aceptacion:
        tiempo_produccion = produccion.fecha_finalizacion - produccion.fecha_aceptacion
    
    context = {
        'produccion': produccion,
        'recetas_produccion': recetas_produccion,
        'ingredientes_necesarios': ingredientes_necesarios,
        'tiempo_transcurrido': tiempo_transcurrido,
        'tiempo_produccion': tiempo_produccion,
        'puede_cambiar_estado': puede_cambiar_estado_produccion(produccion),
    }
    
    return render(request, 'produccion/detalle.html', context)


def eliminar_produccion(request, produccion_id):
    """
    Elimina una producción
    Equivalente a: produccionBean.java - eliminar()
    """
    try:
        produccion = get_object_or_404(Produccion, id_proc=produccion_id)
        
        # Solo se puede eliminar si está pendiente
        if produccion.estado not in ['Pendiente', 'Esperando insumos']:
            messages.error(request, 'Solo se pueden eliminar producciones pendientes o esperando insumos')
            return redirect('produccion:index')
        
        produccion_num = produccion.id_proc
        produccion.delete()
        
        messages.success(request, f'Producción #{produccion_num} eliminada correctamente')
        
    except Exception as e:
        messages.error(request, f'Error al eliminar la producción: {str(e)}')
    
    return redirect('produccion:index')


# Funciones auxiliares

def validar_stock_produccion(receta, cantidad):
    """
    Valida si hay stock suficiente para una producción
    Equivalente a: produccionBean.java - validarStockProduccion()
    """
    try:
        ingredientes = RecetaInsumo.objects.filter(receta=receta)
        
        for ingrediente in ingredientes:
            cantidad_necesaria = ingrediente.cantidad * cantidad
            stock_disponible = ingrediente.insumo.stock_actual
            
            if stock_disponible < cantidad_necesaria:
                return False, f"No hay stock suficiente para {ingrediente.insumo.nombre} (disponible: {stock_disponible}, necesario: {cantidad_necesaria})"
        
        return True, "Stock suficiente"
    
    except Exception as e:
        return False, f"Error validando stock: {str(e)}"


def descontar_stock_produccion(produccion):
    """
    Descuenta el stock al finalizar una producción
    Equivalente a: produccionBean.java - descontarStockProduccion()
    """
    try:
        ingredientes = RecetaInsumo.objects.filter(receta=produccion.receta)
        
        for ingrediente in ingredientes:
            cantidad_usar = ingrediente.cantidad * produccion.cantidad
            
            # Descontar del stock general
            insumo = ingrediente.insumo
            insumo.stock_actual -= cantidad_usar
            insumo.save()
            
            # Descontar de lotes específicos
            DetalleInsumo.descontar_de_lotes(insumo, cantidad_usar)
        
        return True
    
    except Exception as e:
        print(f"Error descontando stock: {e}")
        return False


def marcar_venta_completada(produccion):
    """
    Marca la venta como completada si existe
    Equivalente a: produccionBean.java - marcar venta completada
    """
    try:
        # Buscar si hay una venta asociada a esta producción
        venta_produccion = VentaProduccion.objects.filter(
            produccion=produccion
        ).first()
        
        if venta_produccion:
            venta = venta_produccion.venta
            venta.estado = 'Completada'
            venta.save()
    
    except Exception as e:
        print(f"Error marcando venta como completada: {e}")


def puede_cambiar_estado_produccion(produccion):
    """
    Verifica si se puede cambiar el estado de una producción
    """
    estados_cambiables = ['Pendiente', 'En Proceso', 'Esperando insumos']
    return produccion.estado in estados_cambiables


# API endpoints para AJAX

def api_empleados_produccion(request):
    """
    API para obtener empleados de producción
    """
    empleados = Usuario.objects.filter(
        rol='EP', estado='Activo'
    ).values('id_usu', 'nombres', 'apellidos')
    
    empleados_list = [
        {
            'id': emp['id_usu'],
            'nombre': f"{emp['nombres']} {emp['apellidos']}"
        }
        for emp in empleados
    ]
    
    return JsonResponse({'empleados': empleados_list})


def api_recetas_activas(request):
    """
    API para obtener recetas activas
    """
    recetas = Receta.objects.filter(
        estado='Activa'
    ).values('id_rec', 'nombre', 'precio')
    
    recetas_list = [
        {
            'id': rec['id_rec'],
            'nombre': rec['nombre'],
            'precio': float(rec['precio'])
        }
        for rec in recetas
    ]
    
    return JsonResponse({'recetas': recetas_list})


def api_validar_stock(request):
    """
    API para validar stock antes de crear producción
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            receta_id = data.get('receta_id')
            cantidad = Decimal(str(data.get('cantidad', 0)))
            
            receta = get_object_or_404(Receta, id_rec=receta_id)
            tiene_stock, mensaje = validar_stock_produccion(receta, cantidad)
            
            return JsonResponse({
                'tiene_stock': tiene_stock,
                'mensaje': mensaje
            })
        
        except Exception as e:
            return JsonResponse({
                'tiene_stock': False,
                'mensaje': f'Error validando stock: {str(e)}'
            }, status=400)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def obtener_fechas_produccion(request, produccion_id):
    """
    API para obtener las fechas de trazabilidad de una producción
    """
    try:
        produccion = get_object_or_404(Produccion, id_proc=produccion_id)
        
        return JsonResponse({
            'success': True,
            'fecha_hora': produccion.fecha_hora.strftime('%d/%m/%Y %H:%M:%S') if produccion.fecha_hora else None,
            'fecha_aceptacion': produccion.fecha_aceptacion.strftime('%d/%m/%Y %H:%M:%S') if produccion.fecha_aceptacion else None,
            'fecha_finalizacion': produccion.fecha_finalizacion.strftime('%d/%m/%Y %H:%M:%S') if produccion.fecha_finalizacion else None,
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener fechas: {str(e)}'
        }, status=400)


@login_required  
def exportar_excel(request):
    """
    Exporta las producciones a Excel
    """
    try:
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Crear libro de trabajo
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Producciones"
        
        # Encabezados
        headers = ['ID', 'Fecha Creación', 'Estado', 'Usuario Creador', 'Usuario Asignado', 'Fecha Aceptación', 'Fecha Finalización']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Datos
        producciones = Produccion.objects.all().order_by('-fecha_hora')
        for row, prod in enumerate(producciones, 2):
            ws.cell(row=row, column=1, value=prod.id_proc)
            ws.cell(row=row, column=2, value=prod.fecha_hora.strftime('%d/%m/%Y %H:%M:%S') if prod.fecha_hora else '')
            ws.cell(row=row, column=3, value=prod.estado)
            ws.cell(row=row, column=4, value=f"{prod.usuario.nombre} {prod.usuario.apellido}" if prod.usuario else '')
            ws.cell(row=row, column=5, value=f"{prod.asignado_a.nombre} {prod.asignado_a.apellido}" if prod.asignado_a else 'Sin asignar')
            ws.cell(row=row, column=6, value=prod.fecha_aceptacion.strftime('%d/%m/%Y %H:%M:%S') if prod.fecha_aceptacion else 'Pendiente')
            ws.cell(row=row, column=7, value=prod.fecha_finalizacion.strftime('%d/%m/%Y %H:%M:%S') if prod.fecha_finalizacion else 'Pendiente')
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="producciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'Error: openpyxl no está instalado')
        return redirect('produccion:index')
    except Exception as e:
        messages.error(request, f'Error al exportar Excel: {str(e)}')
        return redirect('produccion:index')


@login_required
def reporte_pdf(request):
    """
    Genera reporte PDF de producciones
    """
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from io import BytesIO
        
        # Crear buffer
        buffer = BytesIO()
        
        # Crear documento PDF
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        normal_style = styles['Normal']
        
        # Título
        title = Paragraph("Reporte de Producciones - PAE", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Fecha del reporte
        fecha_reporte = Paragraph(f"Fecha del reporte: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", normal_style)
        story.append(fecha_reporte)
        story.append(Spacer(1, 20))
        
        # Datos de la tabla
        producciones = Produccion.objects.all().order_by('-fecha_hora')
        data = [['ID', 'Fecha', 'Estado', 'Creador', 'Asignado']]
        
        for prod in producciones:
            data.append([
                str(prod.id_proc),
                prod.fecha_hora.strftime('%d/%m/%Y') if prod.fecha_hora else '',
                prod.estado,
                f"{prod.usuario.nombre} {prod.usuario.apellido}" if prod.usuario else '',
                f"{prod.asignado_a.nombre} {prod.asignado_a.apellido}" if prod.asignado_a else 'Sin asignar'
            ])
        
        # Crear tabla
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Construir PDF
        doc.build(story)
        
        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_producciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        return response
        
    except ImportError:
        messages.error(request, 'Error: reportlab no está instalado')
        return redirect('produccion:index')
    except Exception as e:
        messages.error(request, f'Error al generar PDF: {str(e)}')
        return redirect('produccion:index')
