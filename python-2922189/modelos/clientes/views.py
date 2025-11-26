from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.db import transaction
from decimal import Decimal
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
import logging

from .models import Cliente

logger = logging.getLogger(__name__)


# ==================== LISTAR CLIENTES ====================
def index(request):
    """
    Vista principal de clientes - Equivalente a clientesBean.init()
    Inicializa y carga la lista de clientes con filtros
    """
    try:
        logger.info("Iniciando vista principal de clientes")
        return listar_clientes(request)
    except Exception as e:
        logger.error(f"Error en vista principal de clientes: {str(e)}")
        messages.error(request, f'❌ Error al cargar clientes: {str(e)}')
        return render(request, 'clientes/index.html', {'clientes': [], 'total_clientes': 0})


def listar_clientes(request):
    """
    Lista todos los clientes con filtros - Equivalente a clientesBean.listar() y cargarClientes()
    Implementa la misma lógica de filtrado que clientesDao.filtrar(filtroId, filtroNombre, filtroTelefono, filtroCorreo)
    """
    try:
        # Obtener parámetros de filtro (igual que en clientesBean)
        filtro_id = request.GET.get('filtro_id', '').strip()
        filtro_nombre = request.GET.get('filtro_nombre', '').strip()
        filtro_telefono = request.GET.get('filtro_telefono', '').strip()
        filtro_correo = request.GET.get('filtro_correo', '').strip()
        
        logger.info(f"Aplicando filtros - ID: {filtro_id}, Nombre: {filtro_nombre}, Teléfono: {filtro_telefono}, Correo: {filtro_correo}")
        
        # Query base - equivalente a dao.listar()
        clientes = Cliente.objects.all()
        
        # Aplicar filtros exactamente como en clientesDao.filtrar()
        if filtro_id:
            try:
                id_cliente = int(filtro_id)
                clientes = clientes.filter(id_Cliente=id_cliente)
            except ValueError:
                messages.warning(request, '⚠️ ID de cliente debe ser un número')
                clientes = clientes.none()
        
        if filtro_nombre:
            clientes = clientes.filter(nombre__icontains=filtro_nombre)
        
        if filtro_telefono:
            clientes = clientes.filter(telefono__icontains=filtro_telefono)
        
        if filtro_correo:
            clientes = clientes.filter(correo__icontains=filtro_correo)
        
        # Contar total antes de paginación
        total_clientes = Cliente.objects.count()
        clientes_filtrados = clientes.count()
        
        # Ordenar por nombre como en el bean
        clientes = clientes.order_by('nombre')
        
        # Paginación
        paginator = Paginator(clientes, 10)  # 10 clientes por página
        page = request.GET.get('page', 1)
        clientes_paginados = paginator.get_page(page)
        
        logger.info(f"Clientes encontrados: {clientes_filtrados} de {total_clientes} total")
        
        context = {
            'clientes': clientes_paginados,
            'total_clientes': total_clientes,
            'clientes_filtrados': clientes_filtrados,
            'filtro_id': filtro_id,
            'filtro_nombre': filtro_nombre,
            'filtro_telefono': filtro_telefono,
            'filtro_correo': filtro_correo,
            'hay_filtros': bool(filtro_id or filtro_nombre or filtro_telefono or filtro_correo),
        }
        
        return render(request, 'clientes/index.html', context)
        
    except Exception as e:
        logger.error(f"Error al listar clientes: {str(e)}")
        messages.error(request, f'❌ Error al cargar la lista de clientes: {str(e)}')
        return render(request, 'clientes/index.html', {
            'clientes': [],
            'total_clientes': 0,
            'clientes_filtrados': 0,
            'filtro_id': '',
            'filtro_nombre': '',
            'filtro_telefono': '',
            'filtro_correo': '',
            'hay_filtros': False,
        })


# ==================== NUEVO CLIENTE ====================
def nuevo_cliente(request):
    """
    Vista para crear un nuevo cliente - Equivalente a clientesBean.prepararNuevoCliente()
    Prepara el formulario para un nuevo cliente
    """
    if request.method == 'POST':
        return guardar_cliente(request)
    
    # Preparar cliente nuevo (equivalente a clienteNuevo = new clientes())
    context = {
        'cliente_nuevo': True,
        'titulo': 'Registrar Nuevo Cliente'
    }
    
    return render(request, 'clientes/nuevo.html', context)


@transaction.atomic
def guardar_cliente(request):
    """
    Guarda un nuevo cliente - Equivalente a clientesBean.guardarCliente()
    Implementa la misma lógica de validación y guardado que dao.agregar()
    """
    try:
        logger.info("Iniciando proceso de guardado de cliente")
        
        # Extraer datos del formulario
        nombre = request.POST.get('nombre', '').strip()
        nit = request.POST.get('nit', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        correo = request.POST.get('correo', '').strip()
        
        logger.info(f"Datos recibidos - Nombre: {nombre}, NIT: {nit}, Teléfono: {telefono}, Correo: {correo}")
        
        # Validaciones obligatorias (como en el bean)
        if not nombre:
            raise ValueError('El nombre del cliente es obligatorio')
        
        if not nit:
            raise ValueError('El NIT del cliente es obligatorio')
        
        if not telefono:
            raise ValueError('El teléfono del cliente es obligatorio')
        
        if not correo:
            raise ValueError('El correo del cliente es obligatorio')
        
        # Validar formato de correo básico
        if '@' not in correo or '.' not in correo:
            raise ValueError('El correo electrónico no tiene un formato válido')
        
        # Verificar duplicados (como validaría la base de datos en Java)
        if Cliente.objects.filter(nit=nit).exists():
            raise ValueError(f'Ya existe un cliente registrado con el NIT: {nit}')
        
        if Cliente.objects.filter(telefono=telefono).exists():
            raise ValueError(f'Ya existe un cliente registrado con el teléfono: {telefono}')
        
        if Cliente.objects.filter(correo__iexact=correo).exists():
            raise ValueError(f'Ya existe un cliente registrado con el correo: {correo}')
        
        # Crear y guardar cliente (equivalente a dao.agregar())
        cliente = Cliente(
            nombre=nombre,
            nit=nit,
            telefono=telefono,
            correo=correo.lower()  # Normalizar correo
        )
        cliente.save()
        
        # Obtener ID generado (como retorna dao.agregar())
        cliente_id = cliente.id_Cliente
        
        logger.info(f"Cliente creado exitosamente con ID: {cliente_id}")
        
        # Mensaje de éxito (como en clientesBean)
        messages.success(request, f'✅ Cliente "{nombre}" registrado correctamente con ID #{cliente_id}')
        
        # Redireccionar a la lista (como en clientesBean que retorna a Index.xhtml)
        return redirect('clientes:index')
        
    except ValueError as ve:
        # Errores de validación
        logger.warning(f"Error de validación al crear cliente: {str(ve)}")
        messages.error(request, str(ve))
        return render(request, 'clientes/nuevo.html', {
            'nombre': request.POST.get('nombre', ''),
            'nit': request.POST.get('nit', ''),
            'telefono': request.POST.get('telefono', ''),
            'correo': request.POST.get('correo', ''),
            'cliente_nuevo': True,
            'titulo': 'Registrar Nuevo Cliente'
        })
        
    except Exception as e:
        # Errores generales
        logger.error(f"Error inesperado al crear cliente: {str(e)}")
        messages.error(request, f'❌ Error inesperado al registrar el cliente: {str(e)}')
        return render(request, 'clientes/nuevo.html', {
            'nombre': request.POST.get('nombre', ''),
            'nit': request.POST.get('nit', ''),
            'telefono': request.POST.get('telefono', ''),
            'correo': request.POST.get('correo', ''),
            'cliente_nuevo': True,
            'titulo': 'Registrar Nuevo Cliente'
        })


# ==================== EDITAR CLIENTE ====================
def editar_cliente(request, cliente_id):
    """
    Vista para editar un cliente - Equivalente a clientesBean.prepararEdicion()
    Prepara el cliente seleccionado para edición
    """
    try:
        # Obtener cliente seleccionado (equivalente a clienteSeleccionado = c)
        cliente = get_object_or_404(Cliente, id_Cliente=cliente_id)
        
        logger.info(f"Preparando edición del cliente ID: {cliente_id} - {cliente.nombre}")
        
        if request.method == 'POST':
            return actualizar_cliente(request, cliente)
        
        context = {
            'cliente': cliente,
            'cliente_nuevo': False,
            'titulo': f'Editar Cliente - {cliente.nombre}'
        }
        
        return render(request, 'clientes/editar.html', context)
        
    except Exception as e:
        logger.error(f"Error al preparar edición del cliente {cliente_id}: {str(e)}")
        messages.error(request, f'❌ Error al cargar el cliente para edición: {str(e)}')
        return redirect('clientes:index')


@transaction.atomic
def actualizar_cliente(request, cliente):
    """
    Actualiza un cliente existente - Equivalente a clientesBean.actualizarCliente()
    Implementa la misma lógica que dao.actualizarEnCascada()
    """
    try:
        logger.info(f"Iniciando actualización del cliente ID: {cliente.id_Cliente}")
        
        # Extraer datos del formulario
        nombre = request.POST.get('nombre', '').strip()
        nit = request.POST.get('nit', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        correo = request.POST.get('correo', '').strip()
        
        logger.info(f"Datos para actualizar - Nombre: {nombre}, NIT: {nit}, Teléfono: {telefono}, Correo: {correo}")
        
        # Validaciones obligatorias
        if not nombre:
            raise ValueError('El nombre del cliente es obligatorio')
        
        if not nit:
            raise ValueError('El NIT del cliente es obligatorio')
        
        if not telefono:
            raise ValueError('El teléfono del cliente es obligatorio')
        
        if not correo:
            raise ValueError('El correo del cliente es obligatorio')
        
        # Validar formato de correo
        if '@' not in correo or '.' not in correo:
            raise ValueError('El correo electrónico no tiene un formato válido')
        
        # Verificar duplicados excluyendo el cliente actual (como en el bean)
        if Cliente.objects.filter(nit=nit).exclude(id_Cliente=cliente.id_Cliente).exists():
            raise ValueError(f'Ya existe otro cliente registrado con el NIT: {nit}')
        
        if Cliente.objects.filter(telefono=telefono).exclude(id_Cliente=cliente.id_Cliente).exists():
            raise ValueError(f'Ya existe otro cliente registrado con el teléfono: {telefono}')
        
        if Cliente.objects.filter(correo__iexact=correo).exclude(id_Cliente=cliente.id_Cliente).exists():
            raise ValueError(f'Ya existe otro cliente registrado con el correo: {correo}')
        
        # Guardar datos anteriores para el log
        datos_anteriores = f"{cliente.nombre} - {cliente.nit} - {cliente.telefono} - {cliente.correo}"
        
        # Actualizar cliente (equivalente a dao.actualizarEnCascada())
        cliente.nombre = nombre
        cliente.nit = nit
        cliente.telefono = telefono
        cliente.correo = correo.lower()  # Normalizar correo
        cliente.save()
        
        logger.info(f"Cliente actualizado exitosamente. Datos anteriores: {datos_anteriores}")
        
        # Mensaje de éxito (como en clientesBean)
        messages.success(request, f'✅ Cliente "{nombre}" actualizado correctamente')
        
        # Redireccionar a la lista (como retorna a Index.xhtml)
        return redirect('clientes:index')
        
    except ValueError as ve:
        # Errores de validación
        logger.warning(f"Error de validación al actualizar cliente {cliente.id_Cliente}: {str(ve)}")
        messages.error(request, str(ve))
        return render(request, 'clientes/editar.html', {
            'cliente': cliente,
            'cliente_nuevo': False,
            'titulo': f'Editar Cliente - {cliente.nombre}',
            'nombre': request.POST.get('nombre', cliente.nombre),
            'nit': request.POST.get('nit', cliente.nit),
            'telefono': request.POST.get('telefono', cliente.telefono),
            'correo': request.POST.get('correo', cliente.correo),
        })
        
    except Exception as e:
        # Errores generales
        logger.error(f"Error inesperado al actualizar cliente {cliente.id_Cliente}: {str(e)}")
        messages.error(request, f'❌ Error inesperado al actualizar el cliente: {str(e)}')
        return redirect('clientes:editar', cliente_id=cliente.id_Cliente)


# ==================== ELIMINAR CLIENTE ====================
@transaction.atomic
def eliminar_cliente(request, cliente_id):
    """
    Elimina un cliente - Equivalente a clientesBean.eliminarCliente() y eliminarClienteSeleccionado()
    Implementa la misma lógica que dao.eliminarEnCascada()
    """
    if request.method != 'POST':
        messages.error(request, '❌ Método no permitido para eliminar cliente')
        return redirect('clientes:index')
    
    try:
        logger.info(f"Iniciando eliminación del cliente ID: {cliente_id}")
        
        # Obtener cliente seleccionado para eliminar (como clienteSeleccionadoParaEliminar)
        cliente = get_object_or_404(Cliente, id_Cliente=cliente_id)
        nombre_cliente = cliente.nombre
        
        logger.info(f"Cliente seleccionado para eliminar: {cliente_id} - {nombre_cliente}")
        
        # Verificar si tiene relaciones que impidan la eliminación (como dao.eliminarEnCascada())
        # Verificar ventas asociadas
        from modelos.ventas.models import Venta
        ventas_asociadas = Venta.objects.filter(cliente=cliente).count()
        
        if ventas_asociadas > 0:
            raise ValueError(f'No se puede eliminar el cliente "{nombre_cliente}" porque tiene {ventas_asociadas} venta(s) asociada(s)')
        
        # Aquí podrían agregarse más verificaciones de integridad referencial si es necesario
        # Por ejemplo: pedidos, facturas, etc.
        
        # Proceder con la eliminación (equivalente a dao.eliminarEnCascada())
        cliente.delete()
        
        logger.info(f"Cliente eliminado exitosamente: {nombre_cliente}")
        
        # Actualizar lista (equivalente a listar() en el bean)
        # El redirect hace que se recargue la lista automáticamente
        
        # Mensaje de éxito (como en clientesBean)
        messages.success(request, f'✅ Cliente "{nombre_cliente}" eliminado correctamente')
        
        # Soporte para AJAX
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True, 
                'mensaje': f'Cliente "{nombre_cliente}" eliminado correctamente',
                'cliente_id': cliente_id
            })
        
        return redirect('clientes:index')
        
    except ValueError as ve:
        # Errores de validación de negocio
        logger.warning(f"Error de validación al eliminar cliente {cliente_id}: {str(ve)}")
        error_msg = str(ve)
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg})
        
        messages.error(request, error_msg)
        return redirect('clientes:index')
        
    except Exception as e:
        # Errores generales
        logger.error(f"Error inesperado al eliminar cliente {cliente_id}: {str(e)}")
        error_msg = f'❌ Error inesperado al eliminar el cliente: {str(e)}'
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg})
        
        messages.error(request, error_msg)
        return redirect('clientes:index')


# ==================== EXPORTAR A EXCEL ====================
def exportar_excel(request):
    """
    Exporta la lista de clientes a Excel - Equivalente al proceso inverso de clientesBean.migrar()
    Genera un archivo Excel con el formato esperado para importación
    """
    try:
        logger.info("Iniciando exportación de clientes a Excel")
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes PAE"
        
        # Encabezados (mismo formato que espera la importación)
        headers = ['Nombre', 'Teléfono', 'Correo', 'NIT']  # Orden como en migrar()
        ws.append(headers)
        
        # Aplicar formato a encabezados (colores PAE)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="b71c1c", end_color="b71c1c", fill_type="solid")
        
        # Obtener todos los clientes ordenados por nombre
        clientes = Cliente.objects.all().order_by('nombre')
        total_clientes = clientes.count()
        
        logger.info(f"Exportando {total_clientes} clientes")
        
        # Agregar datos en el mismo orden que la importación
        for cliente in clientes:
            ws.append([
                cliente.nombre,      # Columna A - como en migrar()
                cliente.telefono,    # Columna B - como en migrar()
                cliente.correo,      # Columna C - como en migrar()
                cliente.nit         # Columna D - agregado para completitud
            ])
        
        # Ajustar ancho de columnas para mejor visualización
        ws.column_dimensions['A'].width = 35  # Nombre
        ws.column_dimensions['B'].width = 15  # Teléfono
        ws.column_dimensions['C'].width = 30  # Correo
        ws.column_dimensions['D'].width = 15  # NIT
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nombre del archivo con timestamp
        from datetime import datetime
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Clientes_PAE_{fecha_actual}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        logger.info(f"Exportación completada: {filename} con {total_clientes} clientes")
        
        return response
        
    except Exception as e:
        logger.error(f"Error al exportar clientes a Excel: {str(e)}")
        messages.error(request, f'❌ Error al exportar clientes a Excel: {str(e)}')
        return redirect('clientes:index')


# ==================== IMPORTAR DESDE EXCEL ====================
@transaction.atomic
def importar_excel(request):
    """
    Importa clientes desde un archivo Excel - Equivalente a clientesBean.migrar()
    Implementa la misma lógica de procesamiento de archivos Excel
    """
    if request.method != 'POST':
        messages.error(request, '❌ Método no permitido para importar')
        return redirect('clientes:index')
    
    try:
        logger.info("Iniciando proceso de importación de clientes desde Excel")
        
        # Obtener archivo (equivalente a Part excel en el bean)
        archivo = request.FILES.get('excel')
        
        if not archivo:
            raise ValueError('No se ha seleccionado ningún archivo Excel')
        
        if not archivo.name.endswith(('.xlsx', '.xls')):
            raise ValueError('El archivo debe ser de tipo Excel (.xlsx o .xls)')
        
        logger.info(f"Procesando archivo: {archivo.name}")
        
        # Cargar workbook (equivalente a WorkbookFactory.create())
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active  # Equivalente a libro.getSheetAt(0)
        
        clientes_importados = 0
        errores = []
        filas_procesadas = 0
        
        # Iterar filas saltando cabecera (como en el bean: itrFila.next())
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                filas_procesadas += 1
                
                # Extraer datos de las columnas (como en el bean)
                # Columna A: Nombre, B: Teléfono, C: Correo, D: NIT
                nombre = str(row[0]).strip() if row[0] else None
                telefono = str(row[1]).strip() if row[1] else None
                correo = str(row[2]).strip() if row[2] else None
                nit = str(row[3]).strip() if row[3] and row[3] != 'None' else None
                
                # Validar datos mínimos (como en el bean: cli.getNombre() != null)
                if not nombre or nombre == 'None':
                    logger.debug(f"Fila {idx}: Nombre vacío, saltando")
                    continue
                
                # NIT es obligatorio para evitar duplicados
                if not nit or nit == 'None':
                    errores.append(f'Fila {idx}: NIT vacío - Cliente: {nombre}')
                    continue
                
                # Valores por defecto si faltan datos opcionales
                if not telefono or telefono == 'None':
                    telefono = '0000000000'
                    logger.debug(f"Fila {idx}: Teléfono vacío, usando valor por defecto")
                
                if not correo or correo == 'None' or '@' not in correo:
                    correo = f'cliente{idx}@pae.com'
                    logger.debug(f"Fila {idx}: Correo vacío o inválido, usando valor por defecto")
                
                # Verificar duplicados antes de crear (importante para integridad)
                if Cliente.objects.filter(nit=nit).exists():
                    errores.append(f'Fila {idx}: Ya existe cliente con NIT {nit}')
                    continue
                
                if Cliente.objects.filter(correo__iexact=correo).exists():
                    errores.append(f'Fila {idx}: Ya existe cliente con correo {correo}')
                    continue
                
                # Crear cliente (equivalente a dao.agregar(cli))
                cliente = Cliente(
                    nombre=nombre,
                    nit=nit,
                    telefono=telefono,
                    correo=correo.lower()
                )
                cliente.save()
                clientes_importados += 1
                
                logger.debug(f"Cliente importado: {nombre} - NIT: {nit}")
                
            except Exception as e:
                error_msg = f'Fila {idx}: Error procesando datos - {str(e)}'
                errores.append(error_msg)
                logger.warning(error_msg)
        
        logger.info(f"Importación completada: {clientes_importados}/{filas_procesadas} clientes importados")
        
        # Mensajes de resultado (como en el bean)
        if clientes_importados > 0:
            messages.success(request, f'✅ Éxito: {clientes_importados} cliente(s) migrado(s) exitosamente')
            
            # Actualizar lista (equivalente a listar() en el bean)
            # El redirect automáticamente recarga la lista
        
        if errores:
            # Mostrar máximo 5 errores para no sobrecargar la interfaz
            for error in errores[:5]:
                messages.warning(request, error)
            
            if len(errores) > 5:
                messages.warning(request, f'⚠️ Se encontraron {len(errores) - 5} errores adicionales')
        
        if clientes_importados == 0 and not errores:
            messages.info(request, '📝 No se encontraron datos válidos para importar')
        
        return redirect('clientes:index')
        
    except Exception as e:
        # Error general (como en el bean)
        logger.error(f"Error fatal migrando clientes: {str(e)}")
        messages.error(request, f'❌ Error fatal migrando clientes: {str(e)}')
        return redirect('clientes:index')


# ==================== BÚSQUEDA AJAX ====================
def buscar_clientes_ajax(request):
    """
    Búsqueda de clientes para autocompletado - Útil para selección en ventas y otros módulos
    Proporciona funcionalidad similar a la búsqueda de clientes en otros formularios
    """
    try:
        termino = request.GET.get('q', '').strip()
        
        if len(termino) < 2:
            return JsonResponse({'clientes': [], 'mensaje': 'Ingrese al menos 2 caracteres'})
        
        logger.debug(f"Búsqueda AJAX de clientes con término: {termino}")
        
        # Búsqueda por múltiples campos (nombre, NIT, teléfono)
        clientes = Cliente.objects.filter(
            Q(nombre__icontains=termino) |
            Q(nit__icontains=termino) |
            Q(telefono__icontains=termino) |
            Q(correo__icontains=termino)
        ).order_by('nombre')[:10]  # Limitar a 10 resultados
        
        resultados = []
        for cliente in clientes:
            resultados.append({
                'id': cliente.id_Cliente,
                'nombre': cliente.nombre,
                'nit': cliente.nit,
                'telefono': cliente.telefono,
                'correo': cliente.correo,
                'texto_completo': f'{cliente.nombre} - {cliente.nit} - {cliente.telefono}'
            })
        
        logger.debug(f"Encontrados {len(resultados)} clientes para término: {termino}")
        
        return JsonResponse({
            'clientes': resultados,
            'total': len(resultados),
            'termino': termino
        })
        
    except Exception as e:
        logger.error(f"Error en búsqueda AJAX de clientes: {str(e)}")
        return JsonResponse({
            'clientes': [],
            'error': f'Error en la búsqueda: {str(e)}',
            'total': 0
        })


# ==================== CARGAR CLIENTES (Para uso en otros módulos) ====================
def cargar_clientes(request):
    """
    Carga todos los clientes activos - Equivalente a clientesBean.cargarClientes()
    Útil para formularios de selección en ventas, pedidos, etc.
    """
    try:
        # Obtener todos los clientes ordenados por nombre
        clientes = Cliente.objects.all().order_by('nombre')
        
        # Si es petición AJAX, devolver JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            clientes_data = []
            for cliente in clientes:
                clientes_data.append({
                    'id': cliente.id_Cliente,
                    'nombre': cliente.nombre,
                    'nit': cliente.nit,
                    'telefono': cliente.telefono,
                    'correo': cliente.correo
                })
            
            return JsonResponse({
                'clientes': clientes_data,
                'total': len(clientes_data)
            })
        
        # Para peticiones normales, devolver contexto
        return render(request, 'clientes/seleccionar.html', {
            'clientes': clientes,
            'total_clientes': clientes.count()
        })
        
    except Exception as e:
        logger.error(f"Error al cargar clientes: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'error': str(e), 'clientes': []})
        
        messages.error(request, f'❌ Error al cargar clientes: {str(e)}')
        return redirect('clientes:index')
